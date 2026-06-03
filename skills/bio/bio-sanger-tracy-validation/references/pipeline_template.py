#!/usr/bin/env python3
"""
Tracy Decompose Sanger 验证流水线 — 通用模板

使用前：按项目实际情况修改下方 CONFIG 字典和 parse_variants()。
每次新增测序批次：在 BATCH_DIRS / integrate_results / OUTPUT_COLUMNS 三处追加。

依赖: openpyxl, samtools (PATH 中), tracy 二进制
"""

import csv
import json
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# ============================================================
# 项目配置  ← 新项目只需修改这里
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent   # 项目根目录

XLSX_PATH   = BASE_DIR / "变异信息.xlsx"             # 主变异信息 xlsx
REF_GENOME  = Path("/path/to/reference.fa")          # samtools faidx 已建索引
TRACY_BIN   = str(Path.home() / "tracy_v0.8.9_macos_arm64")  # ~/tracy_v0.8.9_macos_arm64

# 批次目录：每次新增两行（报告成功 + 报告取消）
BATCH_DIRS = [
    ("batch1",        BASE_DIR / "一代测序第一次验证" / "报告成功"),
    ("batch1_cancel", BASE_DIR / "一代测序第一次验证" / "报告取消"),
    # ("batch2",        BASE_DIR / "一代测序第二次验证" / "报告成功"),
    # ("batch2_cancel", BASE_DIR / "一代测序第二次验证" / "报告取消"),
]

OUT_DIR          = BASE_DIR / "outputs"
TRACY_RESULTS_DIR = OUT_DIR / "tracy_results"
LOCAL_REFS_DIR   = OUT_DIR / "tracy_local_refs"

# 判定参数
FLANK_SIZE           = 500   # 局部参考序列两端各取 N bp
POS_TOLERANCE_BASE   = 5     # 坐标容许偏差基础值(bp)
POS_TOLERANCE_SCALE  = 2     # 额外容许 = indel_len * scale
MIN_QUAL_PASS        = 30    # T/high 质量阈值
MIN_QUAL_MARGINAL    = 10    # T/medium 质量阈值

# ============================================================
# Step 1: 解析变异信息  ← 按实际 xlsx 结构调整列序号
# ============================================================

def parse_variants():
    """从 xlsx 解析变异信息，返回 list of dict。"""
    wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True)

    # Sheet 1: 变异基本信息
    ws = wb["第一次检测"]   # ← 修改为实际 sheet 名
    variants = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[0]
        if not name:
            continue
        variants.append({
            "name":      str(name).strip(),
            "chrom":     str(row[1]).strip(),
            "pos":       int(row[2]),
            "ref":       str(row[5]).strip(),   # ← 调整列序
            "alt":       str(row[6]).strip(),
            "var_type":  str(row[9]).strip()  if row[9]  else "",
            "indel_len": int(row[10])         if row[10] is not None else 0,
            "carrier":   str(row[11]).strip() if row[11] else "",
            "gene":      str(row[20]).strip() if row[20] else "",
            "location":  str(row[23]).strip() if row[23] else "",
        })

    # Sheet 2: 人工判定（可选）
    human_results = {}
    confirmed_set = set()
    if "人工判定" in wb.sheetnames:
        ws2 = wb["人工判定"]
        for i, row in enumerate(ws2.iter_rows(values_only=True)):
            if i == 0 or not row[0]:
                continue
            human_results[str(row[0]).strip()] = {
                "human_result": str(row[1]).strip() if row[1] else "",
                "human_actual": str(row[2]).strip() if row[2] else "",
            }
    if "人工确认真实" in wb.sheetnames:
        ws3 = wb["人工确认真实"]
        for row in ws3.iter_rows(values_only=True):
            if row[0]:
                confirmed_set.add(str(row[0]).strip())
    wb.close()

    for v in variants:
        hr = human_results.get(v["name"], {})
        v["human_result"]    = hr.get("human_result", "")
        v["human_actual"]    = hr.get("human_actual", "")
        v["human_confirmed"] = v["name"] in confirmed_set

    print(f"[Step 1] 解析变异 {len(variants)} 条, "
          f"人工判定 {len(human_results)} 条, "
          f"人工确认真实 {len(confirmed_set)} 条")
    return variants


# ============================================================
# Step 2: 提取局部参考序列
# ============================================================

def extract_local_refs(variants):
    LOCAL_REFS_DIR.mkdir(parents=True, exist_ok=True)
    extracted = skipped = 0
    for v in variants:
        name     = v["name"]
        out_fa   = LOCAL_REFS_DIR / f"{name}.fa"
        if out_fa.exists():
            skipped += 1
            # 缓存 local_ref_start
            header = out_fa.read_text().splitlines()[0]
            m = re.search(r":(\d+)-\d+", header)
            v["local_ref_start"] = int(m.group(1)) if m else max(1, v["pos"] - FLANK_SIZE)
            continue

        start = max(1, v["pos"] - FLANK_SIZE)
        end   = v["pos"] + FLANK_SIZE + max(0, len(v["ref"]) - 1)
        region = f"{v['chrom']}:{start}-{end}"
        result = subprocess.run(
            ["samtools", "faidx", str(REF_GENOME), region],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0 or not result.stdout.strip():
            print(f"  [WARN] samtools faidx 失败: {name}")
            continue
        out_fa.write_text(result.stdout)
        v["local_ref_start"] = start
        extracted += 1

    print(f"[Step 2] 局部参考序列: 新提取 {extracted}, 已有跳过 {skipped}")


# ============================================================
# Step 3: 扫描 ab1 文件
# ============================================================

def scan_ab1_files():
    """
    扫描 BATCH_DIRS，建立 variant_name -> ab1 文件列表 映射。
    命名规则: {variant_name}.{variant_name}-{direction}.{id}.xxx.ab1
    方向字段: -F/-FF/-FFF → F；-R/-RR/-RRR → R
    """
    ab1_map = defaultdict(list)
    for batch_name, batch_dir in BATCH_DIRS:
        if not batch_dir.exists():
            print(f"  [WARN] 目录不存在: {batch_dir}")
            continue
        for f in batch_dir.glob("*.ab1"):
            parts = f.name.split(".")
            if len(parts) < 3:
                continue
            variant_name  = parts[0]
            primer_field  = parts[1]
            if   primer_field.endswith("-RR") or primer_field.endswith("-RRR"):
                direction = "R"
            elif primer_field.endswith("-R"):
                direction = "R"
            elif primer_field.endswith("-FF") or primer_field.endswith("-FFF"):
                direction = "F"
            elif primer_field.endswith("-F"):
                direction = "F"
            else:
                direction = "unknown"
            ab1_map[variant_name].append({
                "batch": batch_name, "direction": direction,
                "path": str(f), "filename": f.name,
            })

    total = sum(len(v) for v in ab1_map.values())
    print(f"[Step 3] 扫描到 {total} 个 ab1 文件, 覆盖 {len(ab1_map)} 个变异")
    return dict(ab1_map)


# ============================================================
# Step 4: 批量运行 tracy decompose（带 JSON 缓存）
# ============================================================

def run_tracy_single(ab1_path, local_ref_fa, output_prefix):
    cmd = [TRACY_BIN, "decompose", "-r", str(local_ref_fa), "-v",
           "-o", str(output_prefix), str(ab1_path)]
    try:
        subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    except subprocess.TimeoutExpired:
        return None
    json_path = f"{output_prefix}.json"
    if not os.path.exists(json_path):
        return None
    try:
        with open(json_path) as fh:
            return json.load(fh)
    except (json.JSONDecodeError, IOError):
        return None


def run_tracy_batch(variants, ab1_map):
    TRACY_RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    all_results = []
    total     = sum(len(ab1_map.get(v["name"], [])) for v in variants)
    processed = cached = 0

    for v in variants:
        name      = v["name"]
        local_ref = LOCAL_REFS_DIR / f"{name}.fa"
        if not local_ref.exists():
            continue
        for ab1_info in ab1_map.get(name, []):
            safe_name  = f"{name}_{ab1_info['batch']}_{ab1_info['direction']}"
            out_prefix = TRACY_RESULTS_DIR / safe_name
            json_path  = Path(f"{out_prefix}.json")

            if json_path.exists() and json_path.stat().st_size > 100:
                try:
                    with open(json_path) as fh:
                        data = json.load(fh)
                    cached += 1
                except (json.JSONDecodeError, IOError):
                    data = run_tracy_single(ab1_info["path"], local_ref, out_prefix)
                    processed += 1
            else:
                data = run_tracy_single(ab1_info["path"], local_ref, out_prefix)
                processed += 1

            entry = {
                "name": name, "batch": ab1_info["batch"],
                "direction": ab1_info["direction"],
                "ab1_path": ab1_info["path"],
                "tracy_success": data is not None,
            }
            if data is not None:
                entry.update(parse_tracy_json(data, v))
            all_results.append(entry)

            done = processed + cached
            if done % 50 == 0 or done == total:
                print(f"  [{done}/{total}] 已处理 (新运行 {processed}, 缓存 {cached})")

    print(f"[Step 4] tracy 运行完成: 新运行 {processed}, 缓存 {cached}, 总 {len(all_results)}")
    return all_results


# ============================================================
# Step 5: 解析 tracy JSON + 判定逻辑
# ============================================================

def parse_tracy_json(data, variant_info):
    result = {
        "hetindel":      data.get("hetindel", 0),
        "allele1_frac":  data.get("allele1fraction", 0),
        "allele2_frac":  data.get("allele2fraction", 0),
        "align1_score":  data.get("align1score", 0),
        "local_ref_start": variant_info.get("local_ref_start",
                              max(1, variant_info["pos"] - FLANK_SIZE)),
    }

    local_start      = result["local_ref_start"]
    expected_local_pos = variant_info["pos"] - local_start + 1
    expected_ref     = variant_info["ref"]
    expected_alt     = variant_info["alt"]
    indel_len_abs    = abs(variant_info.get("indel_len", 1)) or 1
    pos_tolerance    = max(POS_TOLERANCE_BASE, indel_len_abs * POS_TOLERANCE_SCALE)

    variants_data = data.get("variants", {})
    columns = variants_data.get("columns", [])
    rows    = variants_data.get("rows", [])

    best_match = None
    best_match_qual = -1
    all_indels = []

    for row in rows:
        if len(row) < len(columns):
            continue
        rd    = dict(zip(columns, row))
        vtype = rd.get("type", "")
        if vtype not in ("Deletion", "Insertion"):
            continue
        indel = {
            "tracy_pos":    rd.get("pos", 0),
            "tracy_ref":    rd.get("ref", ""),
            "tracy_alt":    rd.get("alt", ""),
            "tracy_qual":   rd.get("qual", 0),
            "tracy_filter": rd.get("filter", ""),
            "tracy_gt":     rd.get("genotype", ""),
            "tracy_type":   vtype,
        }
        all_indels.append(indel)

        pos_diff = abs(indel["tracy_pos"] - expected_local_pos)
        if pos_diff > pos_tolerance:
            continue
        if is_equivalent_indel(expected_ref, expected_alt,
                               indel["tracy_ref"], indel["tracy_alt"],
                               variant_info["var_type"]):
            if indel["tracy_qual"] > best_match_qual:
                best_match = {**indel, "pos_diff": pos_diff}
                best_match_qual = indel["tracy_qual"]

    # 宽松匹配（长度差≤1bp）
    if best_match is None:
        for indel in all_indels:
            pos_diff = abs(indel["tracy_pos"] - expected_local_pos)
            if pos_diff > pos_tolerance:
                continue
            if is_near_equivalent_indel(expected_ref, expected_alt,
                                        indel["tracy_ref"], indel["tracy_alt"]):
                if indel["tracy_qual"] > best_match_qual:
                    best_match = {**indel, "pos_diff": pos_diff, "near_match": True}
                    best_match_qual = indel["tracy_qual"]

    result["n_indels_found"] = len(all_indels)
    result["all_indels"]     = all_indels

    if best_match:
        result.update({
            "match_found":   True,
            "match_pos_diff": best_match["pos_diff"],
            "match_qual":    best_match["tracy_qual"],
            "match_filter":  best_match["tracy_filter"],
            "match_ref":     best_match["tracy_ref"],
            "match_alt":     best_match["tracy_alt"],
            "match_gt":      best_match["tracy_gt"],
            "near_match":    best_match.get("near_match", False),
        })
    else:
        result.update({
            "match_found":   False,
            "match_pos_diff": None, "match_qual": None,
            "match_filter":  None,  "match_ref":  None,
            "match_alt":     None,  "match_gt":   None,
            "near_match":    False,
        })
    return result


def is_equivalent_indel(exp_ref, exp_alt, tracy_ref, tracy_alt, var_type):
    """只比较 DEL/INS 方向 + 净长度，兼容 VCF 左对齐差异。"""
    def _classify(r, a):
        if len(r) > len(a): return "DEL", len(r) - len(a)
        if len(a) > len(r): return "INS", len(a) - len(r)
        return None, 0
    et, el = _classify(exp_ref,   exp_alt)
    tt, tl = _classify(tracy_ref, tracy_alt)
    if et is None or tt is None:
        return False
    return et == tt and el == tl


def is_near_equivalent_indel(exp_ref, exp_alt, tracy_ref, tracy_alt):
    """长度差≤1bp 的近似匹配（tracy 合并相邻 indel 时）。"""
    def _classify(r, a):
        if len(r) > len(a): return "DEL", len(r) - len(a)
        if len(a) > len(r): return "INS", len(a) - len(r)
        return None, 0
    et, el = _classify(exp_ref,   exp_alt)
    tt, tl = _classify(tracy_ref, tracy_alt)
    if et is None or tt is None or et != tt:
        return False
    return abs(el - tl) <= 1


def judge_single_ab1(r):
    """单个 ab1 判定，返回 (call, confidence, notes)。"""
    if not r.get("tracy_success"):
        return "R", "low", "tracy 运行失败"

    if r.get("match_found"):
        qual     = r["match_qual"]
        filt     = r["match_filter"]
        pos_diff = r["match_pos_diff"]
        is_near  = r.get("near_match", False)
        near_tag = "(近似匹配,长度差1bp)" if is_near else ""
        if is_near:
            if filt == "PASS" and qual >= MIN_QUAL_PASS:
                return "T?", "medium", f"PASS q={qual} posdiff={pos_diff} {near_tag}"
            return "T?", "low", f"{filt} q={qual} posdiff={pos_diff} {near_tag}"
        else:
            if filt == "PASS" and qual >= MIN_QUAL_PASS:
                return "T", "high",   f"PASS q={qual} posdiff={pos_diff}"
            if filt == "PASS" and qual >= MIN_QUAL_MARGINAL:
                return "T", "medium", f"PASS q={qual}(marginal) posdiff={pos_diff}"
            if qual >= MIN_QUAL_MARGINAL:
                return "T?", "low",   f"{filt} q={qual} posdiff={pos_diff}"
            return "T?", "low", f"{filt} q={qual}(low) posdiff={pos_diff}"

    hetindel = r.get("hetindel", 0)
    n_indels = r.get("n_indels_found", 0)
    if hetindel == 0 and n_indels == 0:
        af1 = r.get("allele1_frac", 0)
        if af1 > 0.85:
            return "F", "high",   f"无 indel, af1={af1:.2f}(纯合WT)"
        return "F", "medium", f"无 indel, af1={af1:.2f}"
    if n_indels > 0:
        summary = "; ".join(
            f"{x['tracy_type']}@{x['tracy_pos']}({x['tracy_ref']}>{x['tracy_alt']},q={x['tracy_qual']})"
            for x in r.get("all_indels", [])[:3]
        )
        return "F?", "low", f"检出其他 indel: {summary}"
    return "?", "low", f"hetindel={hetindel} 但无匹配变异"


def _call_priority(call):
    return {"T": 5, "T?": 4, "F": 3, "F?": 2, "?": 1, "R": 0, "": -1}.get(call, -1)


def combine_calls(calls, variant_name):
    """综合多个 ab1 判定，返回 (final_call, confidence, notes, best_batch)。"""
    if not calls:
        return "R", "low", "无 ab1 数据", ""

    t_calls  = [c for c in calls if c["call"].startswith("T")]
    f_calls  = [c for c in calls if c["call"] == "F"]
    fq_calls = [c for c in calls if c["call"] == "F?"]
    r_calls  = [c for c in calls if c["call"] == "R"]

    best_call_obj = max(calls, key=lambda c: (
        _call_priority(c["call"]),
        c.get("match_qual") or 0,
    ))
    best_batch = best_call_obj["batch"]

    if t_calls:
        high_t = [c for c in t_calls if c["confidence"] == "high"]
        med_t  = [c for c in t_calls if c["confidence"] == "medium"]
        if f_calls:
            if len(high_t) >= 1:
                return "T?", "low", f"T({len(t_calls)})+F({len(f_calls)})冲突; 有高置信T但也有F", best_batch
            return "?", "low", f"T({len(t_calls)})+F({len(f_calls)})冲突", best_batch
        n_t = len(t_calls)
        if len(high_t) >= 2:
            return "T", "high",   f"{n_t}条T(含{len(high_t)}条high)", best_batch
        if len(high_t) == 1:
            conf = "high" if n_t >= 2 else "medium"
            return "T", conf,     f"{n_t}条T(含1条high)", best_batch
        if len(med_t) >= 1:
            return "T", "medium", f"{n_t}条T(含{len(med_t)}条medium)", best_batch
        return "T?", "low",       f"{n_t}条T?(均low)", best_batch

    if f_calls and not fq_calls:
        high_f = [c for c in f_calls if c["confidence"] == "high"]
        if len(high_f) >= 2:
            return "F", "high",   f"{len(f_calls)}条F(含{len(high_f)}条high)", best_batch
        if len(high_f) == 1:
            return "F", "medium", f"{len(f_calls)}条F(含1条high)", best_batch
        return "F", "low",        f"{len(f_calls)}条F(均非high)", best_batch

    if fq_calls:
        return "F?", "low", f"检出非目标 indel({len(fq_calls)}条)", best_batch

    if r_calls and not f_calls:
        return "R", "low", f"全部 tracy 失败({len(r_calls)}条)", best_batch

    t_n = len(t_calls); f_n = len(f_calls); r_n = len(r_calls)
    return "?", "low", f"混合结果(T={t_n},F={f_n},R={r_n})", best_batch


# ============================================================
# Step 5b: 综合判定（整合各批次结果）
# ============================================================

# !! 新增批次时同步修改这两处 !!
_MAIN_BATCHES   = ["batch1"]                    # "报告成功" 批次列表
_CANCEL_BATCHES = ["batch1_cancel"]             # "报告取消" 批次列表

def integrate_results(variants, all_tracy_results):
    by_variant = defaultdict(list)
    for r in all_tracy_results:
        by_variant[r["name"]].append(r)

    final_results = []
    for v in variants:
        name       = v["name"]
        ab1_results = by_variant.get(name, [])
        entry      = {**v}

        for r in ab1_results:
            r["call"], r["confidence"], r["call_notes"] = judge_single_ab1(r)

        by_bd = {}
        for r in ab1_results:
            key = (r["batch"], r["direction"])
            if key not in by_bd or _call_priority(r["call"]) > _call_priority(by_bd[key]["call"]):
                by_bd[key] = r

        # 写入各批次详情列
        all_batch_tags = _MAIN_BATCHES + _CANCEL_BATCHES
        for batch_tag in all_batch_tags:
            for dir_tag in ["F", "R"]:
                key    = (batch_tag, dir_tag)
                r      = by_bd.get(key)
                prefix = f"{batch_tag}_{dir_tag}"
                if r:
                    entry[f"{prefix}_call"]     = r["call"]
                    entry[f"{prefix}_conf"]     = r["confidence"]
                    entry[f"{prefix}_qual"]     = r.get("match_qual", "")
                    entry[f"{prefix}_hetindel"] = r.get("hetindel", "")
                    entry[f"{prefix}_af1"]      = r.get("allele1_frac", "")
                    entry[f"{prefix}_af2"]      = r.get("allele2_frac", "")
                    entry[f"{prefix}_notes"]    = r.get("call_notes", "")
                else:
                    for suffix in ["call","conf","qual","hetindel","af1","af2","notes"]:
                        entry[f"{prefix}_{suffix}"] = ""

        # 综合判定：主批次优先，cancel 仅兜底
        main_calls = [
            by_bd[(bt, dt)]
            for bt in _MAIN_BATCHES for dt in ["F", "R"]
            if (bt, dt) in by_bd
        ]
        if not main_calls:
            main_calls = [
                by_bd[(bt, dt)]
                for bt in _CANCEL_BATCHES for dt in ["F", "R"]
                if (bt, dt) in by_bd
            ]

        entry["tracy_call"], entry["tracy_confidence"], \
        entry["tracy_notes"], entry["best_batch"] = combine_calls(main_calls, name)

        human = entry.get("human_result", "")
        if entry["tracy_call"] and human:
            entry["agree_with_human"] = (
                entry["tracy_call"].rstrip("?*") == human.rstrip("？")
            )
        else:
            entry["agree_with_human"] = ""

        final_results.append(entry)

    print(f"[Step 5] 综合判定完成, {len(final_results)} 条变异")
    return final_results


# ============================================================
# Step 6: 输出列定义  ← 新增批次时在此追加 14 列
# ============================================================

OUTPUT_COLUMNS = [
    "name", "chrom", "pos", "ref", "alt", "var_type", "indel_len",
    "carrier", "gene", "location",
    "tracy_call", "tracy_confidence", "tracy_notes", "best_batch",
    # batch1 -------------------------------------------------------
    "batch1_F_call", "batch1_F_conf", "batch1_F_qual",
    "batch1_F_hetindel", "batch1_F_af1", "batch1_F_af2", "batch1_F_notes",
    "batch1_R_call", "batch1_R_conf", "batch1_R_qual",
    "batch1_R_hetindel", "batch1_R_af1", "batch1_R_af2", "batch1_R_notes",
    # 追加 batch2: 复制上方 14 行，将 batch1 替换为 batch2 ----------
    # "batch2_F_call", "batch2_F_conf", ...
    # ---------------------------------------------------------------
    "human_result", "human_confirmed", "agree_with_human",
]


# ============================================================
# 输出函数
# ============================================================

def write_tsv(final_results, path):
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS,
                                delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(final_results)
    print(f"  -> {path}")


def write_xlsx(final_results, path):
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracy验证结果"

    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(OUTPUT_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font  = hfill and hfont
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    fills = {
        "T":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "F":  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "R":  PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "?":  PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    }
    def _fill(call):
        if not call: return None
        if str(call).startswith("T"): return fills["T"]
        if str(call).startswith("F"): return fills["F"]
        if str(call) == "R":          return fills["R"]
        if "?" in str(call):          return fills["?"]

    call_cols = [i+1 for i, c in enumerate(OUTPUT_COLUMNS) if c.endswith("_call")]

    for ri, row_data in enumerate(final_results, 2):
        for ci, col in enumerate(OUTPUT_COLUMNS, 1):
            ws.cell(row=ri, column=ci, value=row_data.get(col, "")).alignment = \
                Alignment(horizontal="center")
        for ci in call_cols:
            f = _fill(row_data.get(OUTPUT_COLUMNS[ci-1], ""))
            if f:
                ws.cell(row=ri, column=ci).fill = f

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for ci, col in enumerate(OUTPUT_COLUMNS, 1):
        max_len = max(
            len(col),
            max((min(len(str(r.get(col,""))), 40) for r in final_results[:80]), default=0)
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max_len + 2

    # 汇总 sheet
    ws2 = wb.create_sheet("汇总统计")
    _write_summary_sheet(ws2, final_results)

    wb.save(str(path))
    print(f"  -> {path}")


def _write_summary_sheet(ws, results):
    from openpyxl.styles import Font
    bold = Font(bold=True, size=11)
    total = len(results)
    ws.cell(row=1, column=1, value="Tracy 验证结果汇总").font = Font(bold=True, size=14)
    row = 3
    ws.cell(row=row, column=1, value="Tracy 判定分布").font = bold; row += 1
    ws.cell(row=row, column=1, value="判定"); ws.cell(row=row, column=2, value="数量")
    ws.cell(row=row, column=3, value="比例")
    dist = Counter(r["tracy_call"] for r in results)
    for call in ["T","T?","F","F?","R","?"]:
        n = dist.get(call, 0)
        if not n: continue
        row += 1
        ws.cell(row=row,column=1,value=call); ws.cell(row=row,column=2,value=n)
        ws.cell(row=row,column=3,value=f"{n/total*100:.1f}%")


def write_summary_md(final_results, path):
    total     = len(final_results)
    call_dist = Counter(r["tracy_call"] for r in final_results)
    conf_dist = Counter(r["tracy_confidence"] for r in final_results)
    agree     = sum(1 for r in final_results if r.get("agree_with_human") is True)
    disagree  = sum(1 for r in final_results if r.get("agree_with_human") is False)

    lines = [
        "# Tracy Decompose 验证结果汇总\n",
        f"总变异数: {total}\n\n",
        "## Tracy 判定分布\n",
        "| 判定 | 数量 | 比例 |", "|------|------|------|",
    ]
    for call in ["T","T?","F","F?","R","?"]:
        n = call_dist.get(call, 0)
        if n: lines.append(f"| {call} | {n} | {n/total*100:.1f}% |")
    lines += [
        "\n## 置信度分布\n",
        "| 置信度 | 数量 |", "|--------|------|",
        *[f"| {c} | {conf_dist.get(c,0)} |" for c in ["high","medium","low"]],
        "\n## 与人工判定对比\n",
        f"- 一致: {agree} ({agree/max(agree+disagree,1)*100:.1f}%)",
        f"- 不一致: {disagree}",
        f"- 无人工判定: {total-agree-disagree}",
    ]
    Path(path).write_text("\n".join(lines) + "\n")
    print(f"  -> {path}")


# ============================================================
# 主流程
# ============================================================

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("=" * 60)
    print("Tracy Decompose 批量验证流水线")
    print("=" * 60)

    variants = parse_variants()
    extract_local_refs(variants)
    ab1_map  = scan_ab1_files()
    all_tr   = run_tracy_batch(variants, ab1_map)
    results  = integrate_results(variants, all_tr)

    print("\n[Step 6] 输出结果 ...")
    write_tsv(results,   OUT_DIR / "tracy_validation_report.tsv")
    write_xlsx(results,  OUT_DIR / "tracy_validation_report.xlsx")
    write_summary_md(results, OUT_DIR / "tracy_validation_summary.md")

    print("\n" + "=" * 60)
    print("最终统计:")
    for call in ["T","T?","F","F?","R","?"]:
        n = sum(1 for r in results if r["tracy_call"] == call)
        if n: print(f"  {call}: {n}")
    print("=" * 60)


if __name__ == "__main__":
    main()
